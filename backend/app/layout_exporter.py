from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from math import ceil

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import RawOcrLine

LAYOUT_COLUMNS = 64
LAYOUT_ROWS = 96
LAYOUT_SHEET = "版式还原"
DETAIL_SHEET = "OCR明细"


@dataclass(frozen=True)
class OcrBox:
    text: str
    confidence: float
    left: float
    top: float
    right: float
    bottom: float

    @property
    def width(self) -> float:
        return max(1.0, self.right - self.left)

    @property
    def height(self) -> float:
        return max(1.0, self.bottom - self.top)


def _box_from_line(line: RawOcrLine) -> OcrBox | None:
    if not line.box:
        return None
    xs = [point[0] for point in line.box]
    ys = [point[1] for point in line.box]
    return OcrBox(
        text=line.text.strip(),
        confidence=line.confidence,
        left=min(xs),
        top=min(ys),
        right=max(xs),
        bottom=max(ys),
    )


def _layout_bounds(boxes: list[OcrBox], image_width: int, image_height: int) -> tuple[float, float, float, float]:
    if not boxes:
        return 0.0, 0.0, float(image_width), float(image_height)

    padding_x = max(16.0, image_width * 0.025)
    padding_y = max(16.0, image_height * 0.025)
    left = max(0.0, min(box.left for box in boxes) - padding_x)
    top = max(0.0, min(box.top for box in boxes) - padding_y)
    right = min(float(image_width), max(box.right for box in boxes) + padding_x)
    bottom = min(float(image_height), max(box.bottom for box in boxes) + padding_y)
    if right <= left:
        right = left + 1.0
    if bottom <= top:
        bottom = top + 1.0
    return left, top, right, bottom


def _map_to_axis(value: float, source_min: float, source_max: float, target_size: int) -> int:
    ratio = (value - source_min) / (source_max - source_min)
    return min(target_size, max(1, int(ratio * target_size) + 1))


def _span(size: float, source_size: float, target_size: int, minimum: int = 1) -> int:
    return max(minimum, min(target_size, ceil((size / source_size) * target_size)))


def _region_is_free(occupied: set[tuple[int, int]], row: int, column: int, row_span: int, column_span: int) -> bool:
    return all(
        (current_row, current_column) not in occupied
        for current_row in range(row, row + row_span)
        for current_column in range(column, column + column_span)
    )


def _occupy(occupied: set[tuple[int, int]], row: int, column: int, row_span: int, column_span: int) -> None:
    for current_row in range(row, row + row_span):
        for current_column in range(column, column + column_span):
            occupied.add((current_row, current_column))


def _style_region(sheet, row: int, column: int, row_span: int, column_span: int) -> None:
    thin = Side(style="thin", color="C4CDD8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for current_row in range(row, row + row_span):
        for current_column in range(column, column + column_span):
            cell = sheet.cell(row=current_row, column=current_column)
            cell.border = border
            cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")


def _write_layout_sheet(workbook: Workbook, boxes: list[OcrBox], image_width: int, image_height: int) -> None:
    sheet = workbook.active
    sheet.title = LAYOUT_SHEET
    sheet.sheet_view.showGridLines = True
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.freeze_panes = None

    for column_index in range(1, LAYOUT_COLUMNS + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 2.8
    for row_index in range(1, LAYOUT_ROWS + 1):
        sheet.row_dimensions[row_index].height = 14

    left, top, right, bottom = _layout_bounds(boxes, image_width, image_height)
    bound_width = right - left
    bound_height = bottom - top
    occupied: set[tuple[int, int]] = set()

    for box in sorted(boxes, key=lambda item: (item.top, item.left)):
        if not box.text:
            continue
        row = _map_to_axis(box.top, top, bottom, LAYOUT_ROWS)
        column = _map_to_axis(box.left, left, right, LAYOUT_COLUMNS)
        row_span = min(LAYOUT_ROWS - row + 1, _span(box.height, bound_height, LAYOUT_ROWS))
        natural_text_span = max(1, ceil(len(box.text) / 4))
        coordinate_span = _span(box.width, bound_width, LAYOUT_COLUMNS)
        column_span = min(LAYOUT_COLUMNS - column + 1, max(coordinate_span, natural_text_span))

        while row + row_span <= LAYOUT_ROWS and not _region_is_free(occupied, row, column, row_span, column_span):
            row += 1
        if row + row_span > LAYOUT_ROWS:
            continue

        sheet.cell(row=row, column=column, value=box.text)
        sheet.cell(row=row, column=column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row=row, column=column).font = Font(name="Arial", size=10)
        if len(box.text) >= 14:
            sheet.cell(row=row, column=column).font = Font(name="Arial", size=11, bold=len(box.text) >= 18)
        if row_span > 1 or column_span > 1:
            sheet.merge_cells(start_row=row, start_column=column, end_row=row + row_span - 1, end_column=column + column_span - 1)
        _style_region(sheet, row, column, row_span, column_span)
        _occupy(occupied, row, column, row_span, column_span)


def _write_detail_sheet(workbook: Workbook, boxes: list[OcrBox]) -> None:
    sheet = workbook.create_sheet(DETAIL_SHEET)
    headers = ["序号", "文字", "置信度", "左", "上", "右", "下"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, box in enumerate(boxes, start=1):
        sheet.append([index, box.text, box.confidence, box.left, box.top, box.right, box.bottom])

    widths = [8, 42, 10, 10, 10, 10, 10]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_layout_workbook(
    raw_lines: list[RawOcrLine],
    filename: str,
    image_width: int,
    image_height: int,
) -> bytes:
    boxes = [box for line in raw_lines if (box := _box_from_line(line)) is not None]
    workbook = Workbook()
    workbook.properties.title = f"OCR layout export - {filename}"
    _write_layout_sheet(workbook, boxes, image_width, image_height)
    _write_detail_sheet(workbook, boxes)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
