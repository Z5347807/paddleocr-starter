from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import RecognitionResult, TableData

TITLE = "变压器短路承受能力试验现场记录"
FORM_SHEET = "现场记录"
RAW_SHEET = "OCR明细"
BODY_FONT = "Microsoft YaHei"


def _thin_border() -> Border:
    side = Side(style="thin", color="7F8C9A")
    return Border(left=side, right=side, top=side, bottom=side)


def _medium_border() -> Border:
    side = Side(style="medium", color="4C5967")
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _set(
    sheet,
    cell: str,
    value: object = "",
    *,
    bold: bool = False,
    size: int = 10,
    fill: str = "FFFFFF",
    align: str = "center",
) -> None:
    target = sheet[cell]
    target.value = value
    target.font = Font(name=BODY_FONT, size=size, bold=bold, color="1F2933")
    target.fill = _fill(fill)
    target.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    target.border = _thin_border()


def _style_range(sheet, start_row: int, start_col: int, end_row: int, end_col: int, fill: str = "FFFFFF") -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = _fill(fill)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.font.name is None or cell.font.name == "Calibri":
                cell.font = Font(name=BODY_FONT, size=10, color="1F2933")


def _merge(sheet, range_ref: str, value: object = "", *, bold: bool = False, size: int = 10, fill: str = "FFFFFF") -> None:
    sheet.merge_cells(range_ref)
    anchor = range_ref.split(":", 1)[0]
    _set(sheet, anchor, value, bold=bold, size=size, fill=fill)
    merged = sheet[range_ref]
    rows = [cell.row for row in merged for cell in row]
    cols = [cell.column for row in merged for cell in row]
    _style_range(sheet, min(rows), min(cols), max(rows), max(cols), fill=fill)
    sheet[anchor].value = value
    sheet[anchor].font = Font(name=BODY_FONT, size=size, bold=bold, color="1F2933")


def _field(result: RecognitionResult, field_id: str) -> str:
    return result.fields.get(field_id).value if field_id in result.fields else ""


def _table(result: RecognitionResult, table_id: str) -> TableData | None:
    for table in result.tables:
        if table.id == table_id:
            return table
    return None


def _configure_sheet(sheet) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A7"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35

    widths = {
        "A": 8,
        "B": 8,
        "C": 12,
        "D": 12,
        "E": 12,
        "F": 12,
        "G": 3,
        "H": 3,
        "I": 9,
        "J": 9,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in range(1, 35):
        sheet.row_dimensions[row].height = 24
    sheet.row_dimensions[1].height = 32
    sheet.row_dimensions[29].height = 42


def _write_header(sheet, result: RecognitionResult) -> None:
    header_fill = "EAF1F8"
    label_fill = "F4F7FA"
    _merge(sheet, "A1:N1", TITLE, bold=True, size=16, fill=header_fill)
    _merge(sheet, "L2:L2", "中心编号：", bold=True, fill=label_fill)
    _merge(sheet, "M2:N2", _field(result, "centerNumber"), fill="FFFFFF")

    _merge(sheet, "A3:B3", "委托单位", bold=True, fill=label_fill)
    _merge(sheet, "C3:H3", _field(result, "client"))
    _merge(sheet, "I3:J3", "序号", bold=True, fill=label_fill)
    _merge(sheet, "K3:N3", _field(result, "serialNumber"))

    _merge(sheet, "A4:B4", "型号", bold=True, fill=label_fill)
    _merge(sheet, "C4:H4", _field(result, "model"))
    _merge(sheet, "I4:J4", "联结组", bold=True, fill=label_fill)
    _merge(sheet, "K4:N4", _field(result, "connectionGroup"))

    _merge(sheet, "A5:D5", "样品检验流程卡", fill=label_fill)
    _merge(sheet, "E5:H5", "样品检验标识", fill=label_fill)
    _merge(sheet, "I5:N5", "现场情况", fill=label_fill)


def _write_before_table(sheet, table: TableData | None) -> None:
    section_fill = "DCEAF7"
    _merge(sheet, "A6:F6", "电抗测量（mH）", bold=True, fill=section_fill)
    _merge(sheet, "A7:A10", "试验前", bold=True, fill="F4F7FA")
    headers = ["分接", "AB", "BC", "CA", "电抗(mH)"]
    for index, header in enumerate(headers, start=2):
        _set(sheet, f"{get_column_letter(index)}7", header, bold=True, fill="F4F7FA")

    rows = table.rows if table else []
    for offset in range(3):
        row_data = rows[offset] if offset < len(rows) else {}
        excel_row = 8 + offset
        for index, header in enumerate(headers, start=2):
            _set(sheet, f"{get_column_letter(index)}{excel_row}", row_data.get(header, ""))


def _write_during_table(sheet, table: TableData | None) -> None:
    _merge(sheet, "A12:A22", "试验过程中", bold=True, fill="F4F7FA")
    headers = ["分接", "AB", "BC", "CA", "试验次数"]
    for index, header in enumerate(headers, start=2):
        _set(sheet, f"{get_column_letter(index)}13", header, bold=True, fill="F4F7FA")
    _merge(sheet, "B12:F12", "电抗测量（mH）", bold=True, fill="DCEAF7")

    rows = table.rows if table else []
    for offset in range(9):
        row_data = rows[offset] if offset < len(rows) else {}
        excel_row = 14 + offset
        for index, header in enumerate(headers, start=2):
            _set(sheet, f"{get_column_letter(index)}{excel_row}", row_data.get(header, ""))


def _write_after_table(sheet, table: TableData | None) -> None:
    _merge(sheet, "A25:A27", "试验后", bold=True, fill="F4F7FA")
    headers = ["分接", "AB", "BC", "CA"]
    for index, header in enumerate(headers, start=2):
        _set(sheet, f"{get_column_letter(index)}25", header, bold=True, fill="F4F7FA")

    rows = table.rows if table else []
    for offset in range(2):
        row_data = rows[offset] if offset < len(rows) else {}
        excel_row = 26 + offset
        for index, header in enumerate(headers, start=2):
            _set(sheet, f"{get_column_letter(index)}{excel_row}", row_data.get(header, ""))


def _write_site_panel(sheet, result: RecognitionResult) -> None:
    label_fill = "F4F7FA"
    _merge(sheet, "I6:N6", "现场情况", bold=True, fill="DCEAF7")
    site_rows = [
        (7, "中变", "分接", "14"),
        (8, "并井", "电压", "120 kV"),
        (9, "避雷器", "电压", "110 kV"),
        (10, "", "次数", ""),
    ]
    for row, left, label, value in site_rows:
        _merge(sheet, f"I{row}:J{row}", left, fill=label_fill if left else "FFFFFF")
        _merge(sheet, f"K{row}:L{row}", label, bold=True, fill=label_fill)
        _merge(sheet, f"M{row}:N{row}", value)

    _merge(sheet, "I12:N12", "相别 / 电抗器 / 接地棒 / 电抗测量线", bold=True, fill="DCEAF7")
    right_rows = [
        ("C-AB", "40.0", "√", "√"),
        ("C-AB", "13.5", "√", "√"),
        ("B-AC", "17.0", "√", "√"),
        ("A-BC", "20.5", "√", "√"),
    ]
    headers = ["相别", "电抗器(Ω)", "接地棒", "电抗测量线"]
    for col_offset, header in enumerate(headers, start=9):
        _set(sheet, f"{get_column_letter(col_offset)}13", header, bold=True, fill=label_fill)
    for offset, row_data in enumerate(right_rows, start=14):
        for col_offset, value in enumerate(row_data, start=9):
            _set(sheet, f"{get_column_letter(col_offset)}{offset}", value)

    _merge(sheet, "I20:N20", "继电保护整定", bold=True, fill="DCEAF7")
    relay_rows = [
        ("过流速断", "1.53"),
        ("一次开关延时速断时间", "029"),
        ("二次开关延时速断时间", "125"),
    ]
    for offset, (label, value) in enumerate(relay_rows, start=21):
        _merge(sheet, f"I{offset}:L{offset}", label, bold=True, fill=label_fill)
        _merge(sheet, f"M{offset}:N{offset}", value)

    _merge(sheet, "I25:J25", "备注：", bold=True, fill=label_fill)
    _merge(sheet, "K25:N28", "H-L短路承受能力试验")
    _merge(sheet, "A29:H31", "外观及吊心情况：", fill="FFFFFF")
    _merge(sheet, "A33:B33", "试验：")
    _merge(sheet, "C33:D33", "记录：")
    _merge(sheet, "E33:F33", "校核：")
    _merge(sheet, "K30:L30", "日期：", bold=True, fill=label_fill)
    _merge(sheet, "M30:N30", _field(result, "date"))


def _apply_outer_borders(sheet) -> None:
    ranges = ["A1:N1", "A3:N5", "A6:F10", "A12:F22", "A25:F27", "I6:N10", "I12:L17", "I20:N23", "A29:H31", "I25:N28"]
    border = _medium_border()
    for range_ref in ranges:
        cells = sheet[range_ref]
        rows = [cell.row for row in cells for cell in row]
        cols = [cell.column for row in cells for cell in row]
        min_row, max_row = min(rows), max(rows)
        min_col, max_col = min(cols), max(cols)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if row in {min_row, max_row} or col in {min_col, max_col}:
                    sheet.cell(row=row, column=col).border = border


def _write_form_sheet(workbook: Workbook, result: RecognitionResult) -> None:
    sheet = workbook.active
    sheet.title = FORM_SHEET
    _configure_sheet(sheet)
    _write_header(sheet, result)
    _write_before_table(sheet, _table(result, "before_test_reactance"))
    _write_during_table(sheet, _table(result, "during_test_reactance"))
    _write_after_table(sheet, _table(result, "after_test_reactance"))
    _write_site_panel(sheet, result)
    _apply_outer_borders(sheet)


def _write_raw_sheet(workbook: Workbook, result: RecognitionResult) -> None:
    sheet = workbook.create_sheet(RAW_SHEET)
    headers = ["序号", "文字", "置信度", "左", "上", "右", "下"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(name=BODY_FONT, bold=True)
        cell.fill = _fill("EAF1F8")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    for index, line in enumerate(result.rawOcr, start=1):
        xs = [point[0] for point in line.box] if line.box else [0]
        ys = [point[1] for point in line.box] if line.box else [0]
        sheet.append([index, line.text, line.confidence, min(xs), min(ys), max(xs), max(ys)])

    widths = [8, 42, 10, 12, 12, 12, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_form_workbook(result: RecognitionResult) -> bytes:
    workbook = Workbook()
    workbook.properties.title = TITLE
    _write_form_sheet(workbook, result)
    _write_raw_sheet(workbook, result)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
