from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import RecognitionResult

TITLE = "变压器短路承受能力试验现场记录"


def _style_header(row) -> None:
    fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(sheet) -> None:
    for column in sheet.columns:
        width = 12
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 4, 42))
        sheet.column_dimensions[column_letter].width = width


def _write_result_sheet(workbook: Workbook, result: RecognitionResult) -> None:
    sheet = workbook.active
    sheet.title = "识别结果"
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(bold=True, size=16)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    row_index = 3
    for field in result.fields.values():
        sheet.cell(row=row_index, column=1, value=field.label)
        sheet.cell(row=row_index, column=2, value=field.value)
        row_index += 1

    row_index += 2
    for table in result.tables:
        sheet.cell(row=row_index, column=1, value=table.title)
        sheet.cell(row=row_index, column=1).font = Font(bold=True)
        row_index += 1
        for column_index, column_name in enumerate(table.columns, start=1):
            sheet.cell(row=row_index, column=column_index, value=column_name)
        _style_header(sheet[row_index])
        row_index += 1
        for row in table.rows:
            for column_index, column_name in enumerate(table.columns, start=1):
                sheet.cell(row=row_index, column=column_index, value=row.get(column_name, ""))
            row_index += 1
        row_index += 1

    _autosize(sheet)


def _write_fields_sheet(workbook: Workbook, result: RecognitionResult) -> None:
    sheet = workbook.create_sheet("关键字段")
    headers = ["Field id", "Field label", "Value", "Confidence", "User edited flag"]
    sheet.append(headers)
    _style_header(sheet[1])
    for field_id, field in result.fields.items():
        sheet.append([field_id, field.label, field.value, field.confidence, field.edited])
    _autosize(sheet)


def _write_raw_ocr_sheet(workbook: Workbook, result: RecognitionResult) -> None:
    sheet = workbook.create_sheet("原始OCR")
    headers = ["Text", "Confidence", "Box coordinates"]
    sheet.append(headers)
    _style_header(sheet[1])
    for line in result.rawOcr:
        sheet.append([line.text, line.confidence, str(line.box)])
    _autosize(sheet)


def build_workbook(result: RecognitionResult) -> bytes:
    workbook = Workbook()
    _write_result_sheet(workbook, result)
    _write_fields_sheet(workbook, result)
    _write_raw_ocr_sheet(workbook, result)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
