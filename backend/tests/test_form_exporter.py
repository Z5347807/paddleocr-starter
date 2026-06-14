from io import BytesIO

from openpyxl import load_workbook

from app.form_exporter import build_form_workbook
from app.models import FieldValue, ImageInfo, RawOcrLine, RecognitionResult, TableData


def sample_result() -> RecognitionResult:
    return RecognitionResult(
        documentType="transformer_short_circuit_test_record",
        image=ImageInfo(filename="sample.png", width=1080, height=1920),
        fields={
            "centerNumber": FieldValue(label="中心编号", value="B260112", confidence=0.96),
            "client": FieldValue(label="委托单位", value="山东鲁能泰山电力设备有限公司", confidence=0.95),
            "model": FieldValue(label="型号", value="SSZ22-63000/110", confidence=0.93),
            "serialNumber": FieldValue(label="序号", value="LTD-260-0032", confidence=0.94),
            "connectionGroup": FieldValue(label="联结组", value="YNyn0d11", confidence=0.92),
            "date": FieldValue(label="日期", value="2026年5月25日", confidence=0.9),
        },
        tables=[
            TableData(
                id="before_test_reactance",
                title="试验前电抗测量",
                columns=["分接", "AB", "BC", "CA", "电抗(mH)"],
                rows=[
                    {"分接": "1", "AB": "281.25", "BC": "281.15", "CA": "279.05", "电抗(mH)": "285.348"},
                    {"分接": "9B", "AB": "227.25", "BC": "227.25", "CA": "226.85", "电抗(mH)": "230.322"},
                ],
            ),
            TableData(
                id="during_test_reactance",
                title="试验过程中电抗测量",
                columns=["分接", "AB", "BC", "CA", "试验次数"],
                rows=[
                    {"分接": "1", "AB": "281.45", "BC": "281.45", "CA": "279.45", "试验次数": "1"},
                    {"分接": "9B", "AB": "227.45", "BC": "227.55", "CA": "226.25", "试验次数": "4"},
                ],
            ),
            TableData(
                id="after_test_reactance",
                title="试验后电抗测量",
                columns=["分接", "AB", "BC", "CA"],
                rows=[
                    {"分接": "9B", "AB": "227.45", "BC": "227.45", "CA": "226.25"},
                    {"分接": "1", "AB": "281.45", "BC": "281.45", "CA": "279.45"},
                ],
            ),
        ],
        rawOcr=[
            RawOcrLine(text="中心编号：B260112", confidence=0.96, box=[[1, 2], [3, 2], [3, 4], [1, 4]]),
            RawOcrLine(text="281.25", confidence=0.91, box=[[10, 20], [12, 20], [12, 22], [10, 22]]),
        ],
    )


def test_build_form_workbook_creates_fixed_form_sheet() -> None:
    workbook = load_workbook(BytesIO(build_form_workbook(sample_result())))
    sheet = workbook["现场记录"]

    assert workbook.sheetnames == ["现场记录", "OCR明细"]
    assert sheet["A1"].value == "变压器短路承受能力试验现场记录"
    assert "A1:N1" in [str(merged_range) for merged_range in sheet.merged_cells.ranges]

    assert sheet["A3"].value == "委托单位"
    assert sheet["C3"].value == "山东鲁能泰山电力设备有限公司"
    assert sheet["I3"].value == "序号"
    assert sheet["K3"].value == "LTD-260-0032"
    assert sheet["A4"].value == "型号"
    assert sheet["C4"].value == "SSZ22-63000/110"
    assert sheet["I4"].value == "联结组"
    assert sheet["K4"].value == "YNyn0d11"
    assert sheet["L2"].value == "中心编号："
    assert sheet["M2"].value == "B260112"

    assert sheet["A7"].value == "试验前"
    assert sheet["B7"].value == "分接"
    assert sheet["C8"].value == "281.25"
    assert sheet["F8"].value == "285.348"
    assert sheet["B13"].value == "分接"
    assert sheet["C14"].value == "281.45"
    assert sheet["F15"].value == "4"
    assert sheet["A25"].value == "试验后"
    assert sheet["B26"].value == "9B"
    assert sheet["E27"].value == "279.45"

    assert sheet["A3"].border.left.style in {"thin", "medium"}
    assert sheet["C8"].border.left.style == "thin"
    assert "M30:N30" in [str(merged_range) for merged_range in sheet.merged_cells.ranges]
    assert sheet["M30"].value == "2026年5月25日"


def test_build_form_workbook_keeps_raw_ocr_detail_sheet() -> None:
    workbook = load_workbook(BytesIO(build_form_workbook(sample_result())))
    detail = workbook["OCR明细"]

    assert detail["A1"].value == "序号"
    assert detail["B1"].value == "文字"
    assert detail["B2"].value == "中心编号：B260112"
    assert detail["C2"].value == 0.96
