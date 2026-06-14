from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app, get_engine
from app.models import RawOcrLine

client = TestClient(app)


class FakeEngine:
    def recognize(self, image_bytes: bytes) -> list[RawOcrLine]:
        return [
            RawOcrLine(
                text="中心编号：B260112",
                confidence=0.93,
                box=[[80.0, 40.0], [180.0, 40.0], [180.0, 64.0], [80.0, 64.0]],
            ),
            RawOcrLine(
                text="281.25",
                confidence=0.89,
                box=[[80.0, 160.0], [140.0, 160.0], [140.0, 184.0], [80.0, 184.0]],
            ),
        ]


def test_export_layout_recognizes_image_and_returns_xlsx() -> None:
    app.dependency_overrides[get_engine] = lambda: FakeEngine()
    try:
        response = client.post(
            "/api/export-layout",
            files={"file": ("sample.png", b"not-an-image", "image/png")},
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 400


def test_export_layout_uses_ocr_lines_for_workbook(monkeypatch) -> None:
    from app import main

    app.dependency_overrides[get_engine] = lambda: FakeEngine()
    monkeypatch.setattr(main, "_image_size", lambda image_bytes: (240, 320))
    try:
        response = client.post(
            "/api/export-layout",
            files={"file": ("sample.png", b"fake-image", "image/png")},
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.content))
    assert "版式还原" in workbook.sheetnames
    assert any(cell.value == "中心编号：B260112" for row in workbook["版式还原"].iter_rows() for cell in row)
