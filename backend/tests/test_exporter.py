from io import BytesIO

from openpyxl import load_workbook

from app.exporter import build_workbook
from app.models import FieldValue, ImageInfo, RawOcrLine, RecognitionResult, TableData


def sample_result() -> RecognitionResult:
    return RecognitionResult(
        documentType="transformer_short_circuit_test_record",
        image=ImageInfo(filename="sample.jpg", width=1080, height=1920),
        fields={
            "centerNumber": FieldValue(label="中心编号", value="B260112", confidence=0.93, edited=True),
            "client": FieldValue(label="委托单位", value="山东鲁能泰山电力设备有限公司", confidence=0.91),
        },
        tables=[
            TableData(
                id="before_test_reactance",
                title="试验前电抗测量",
                columns=["分接", "AB", "BC", "CA", "电抗(mH)"],
                rows=[{"分接": "1", "AB": "281.25", "BC": "281.15", "CA": "279.05", "电抗(mH)": "285.348"}],
            )
        ],
        rawOcr=[
            RawOcrLine(
                text="中心编号：B260112",
                confidence=0.93,
                box=[[1.0, 2.0], [3.0, 2.0], [3.0, 4.0], [1.0, 4.0]],
            )
        ],
    )


def test_build_workbook_creates_expected_sheets_and_values() -> None:
    content = build_workbook(sample_result())
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["识别结果", "关键字段", "原始OCR"]
    assert workbook["识别结果"]["A1"].value == "变压器短路承受能力试验现场记录"
    assert workbook["识别结果"]["A3"].value == "中心编号"
    assert workbook["识别结果"]["B3"].value == "B260112"
    assert workbook["识别结果"]["A7"].value == "试验前电抗测量"
    assert workbook["识别结果"]["A9"].value == "1"
    assert workbook["关键字段"]["A2"].value == "centerNumber"
    assert workbook["关键字段"]["E2"].value is True
    assert workbook["原始OCR"]["A2"].value == "中心编号：B260112"
