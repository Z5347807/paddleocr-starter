from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app, get_engine
from app.models import RawOcrLine

client = TestClient(app)


class FakeEngine:
    def recognize(self, image_bytes: bytes) -> list[RawOcrLine]:
        return [
            RawOcrLine(
                text="中心编号：B260112",
                confidence=0.93,
                box=[[80.0, 40.0], [180.0, 40.0], [180.0, 64.0], [80.0, 64.0]],
            ),
            RawOcrLine(
                text="委托单位",
                confidence=0.92,
                box=[[20.0, 80.0], [80.0, 80.0], [80.0, 104.0], [20.0, 104.0]],
            ),
            RawOcrLine(
                text="山东鲁能泰山电力设备有限公司",
                confidence=0.92,
                box=[[120.0, 80.0], [260.0, 80.0], [260.0, 104.0], [120.0, 104.0]],
            ),
            RawOcrLine(
                text="序号",
                confidence=0.91,
                box=[[20.0, 120.0], [80.0, 120.0], [80.0, 144.0], [20.0, 144.0]],
            ),
            RawOcrLine(
                text="LTD-260-0032",
                confidence=0.91,
                box=[[120.0, 120.0], [260.0, 120.0], [260.0, 144.0], [120.0, 144.0]],
            ),
            RawOcrLine(
                text="281.25",
                confidence=0.89,
                box=[[240.0, 632.0], [300.0, 632.0], [300.0, 656.0], [240.0, 656.0]],
            ),
        ]


def test_export_form_rejects_invalid_image() -> None:
    app.dependency_overrides[get_engine] = lambda: FakeEngine()
    try:
        response = client.post(
            "/api/export-form",
            files={"file": ("sample.png", b"not-an-image", "image/png")},
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 400


def test_export_form_uses_ocr_lines_for_fixed_form_workbook(monkeypatch) -> None:
    from app import main

    app.dependency_overrides[get_engine] = lambda: FakeEngine()
    monkeypatch.setattr(main, "_image_size", lambda image_bytes: (240, 320))
    try:
        response = client.post(
            "/api/export-form",
            files={"file": ("sample.png", b"fake-image", "image/png")},
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["现场记录", "OCR明细"]
    assert workbook["现场记录"]["A1"].value == "变压器短路承受能力试验现场记录"
    assert workbook["现场记录"]["M2"].value == "B260112"
    assert workbook["现场记录"]["C3"].value == "山东鲁能泰山电力设备有限公司"
