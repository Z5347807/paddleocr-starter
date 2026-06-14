from io import BytesIO

from openpyxl import load_workbook

from app.layout_exporter import build_layout_workbook
from app.models import RawOcrLine


def ocr_line(text: str, x: float, y: float, w: float = 80.0, h: float = 24.0) -> RawOcrLine:
    return RawOcrLine(
        text=text,
        confidence=0.91,
        box=[
            [x, y],
            [x + w, y],
            [x + w, y + h],
            [x, y + h],
        ],
    )


def find_cell(sheet, text: str) -> tuple[int, int]:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == text:
                return cell.row, cell.column
    raise AssertionError(f"{text!r} was not written to the layout sheet")


def test_build_layout_workbook_places_ocr_text_by_image_position() -> None:
    content = build_layout_workbook(
        raw_lines=[
            ocr_line("变压器短路承受能力试验现场记录", 320, 260, 360, 32),
            ocr_line("委托单位", 80, 360, 88, 26),
            ocr_line("山东鲁能泰山电力设备有限公司", 240, 360, 280, 26),
            ocr_line("中心编号：B260112", 840, 300, 160, 26),
            ocr_line("281.25", 190, 610, 80, 30),
            ocr_line("279.05", 460, 610, 80, 30),
        ],
        filename="sample.png",
        image_width=1080,
        image_height=1920,
    )

    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["版式还原", "OCR明细"]
    layout = workbook["版式还原"]
    detail = workbook["OCR明细"]

    label_row, label_column = find_cell(layout, "委托单位")
    value_row, value_column = find_cell(layout, "山东鲁能泰山电力设备有限公司")
    center_row, center_column = find_cell(layout, "中心编号：B260112")
    ab_row, ab_column = find_cell(layout, "281.25")
    ca_row, ca_column = find_cell(layout, "279.05")

    assert abs(label_row - value_row) <= 1
    assert value_column > label_column
    assert center_row < ab_row
    assert center_column > value_column
    assert ca_column > ab_column

    assert detail.cell(row=1, column=1).value == "序号"
    assert detail.cell(row=2, column=2).value == "变压器短路承受能力试验现场记录"
    assert detail.cell(row=2, column=3).value == 0.91
